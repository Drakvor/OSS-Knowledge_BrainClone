"""
Excel File Processor
Migrated sophisticated logic from original script with new architecture
"""

import os
import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain imports
from langchain_openai import AzureChatOpenAI
from langchain_core.prompts import PromptTemplate, FewShotPromptTemplate
from langchain.output_parsers import PydanticOutputParser
from langchain.output_parsers import OutputFixingParser

# LLMChain is deprecated but still available in 0.2.x for backwards compatibility
try:
    from langchain.chains.llm import LLMChain
except ImportError:
    from langchain.chains import LLMChain

from app.config import settings
from app.processors.base.interfaces import FileProcessorInterface
from app.processors.base.base_models import ProcessedChunk, ChunkRelationship, SemanticAnalysis
from app.processors.excel.excel_models import (
    ExcelProcessingOptions, ExcelFileAnalysis, ExcelSheetAnalysis, 
    ExcelColumnAnalysis, ExcelChunkType, ExcelChunkingStrategy
)
from app.core.exceptions import ProcessingException

logger = logging.getLogger(__name__)


class ExcelSemanticAnalyzer:
    """LangChain-based semantic analyzer for Excel data - migrated from original"""
    
    def __init__(self):
        self.llm = AzureChatOpenAI(
            api_key=settings.OPENAI_API_KEY,
            azure_endpoint=settings.AZURE_OPENAI_ENDPOINT,
            api_version=settings.AZURE_OPENAI_API_VERSION,
            deployment_name=settings.OPENAI_MODEL,
            max_tokens=settings.OPENAI_MAX_TOKENS,
            temperature=settings.OPENAI_TEMPERATURE
        )
        self.setup_output_parsers()
    
    def setup_output_parsers(self):
        """Setup robust output parsers with error handling"""
        self.domain_parser = OutputFixingParser.from_llm(
            parser=PydanticOutputParser(pydantic_object=dict), 
            llm=self.llm
        )
        self.semantic_parser = OutputFixingParser.from_llm(
            parser=PydanticOutputParser(pydantic_object=dict),
            llm=self.llm
        )
        self.relationship_parser = OutputFixingParser.from_llm(
            parser=PydanticOutputParser(pydantic_object=dict),
            llm=self.llm
        )
    
    def create_domain_detection_chain(self) -> LLMChain:
        """Create domain detection chain with few-shot examples - preserved from original"""
        
        examples = [
            {
                "headers": "서비스문의 ID, 등록일시, 상태, 요청자명, 제목",
                "sample_data": "SRM25020387826, 2025-02-03 10:05, 문의 종료, 김민수, M-OSS 설치 경로 문의",
                "domain": "IT서비스관리",
                "confidence": "0.95"
            },
            {
                "headers": "사원번호, 성명, 부서, 직급, 입사일자",
                "sample_data": "EMP001, 김철수, 개발팀, 대리, 2023-03-15",
                "domain": "인사관리",
                "confidence": "0.92"
            },
            {
                "headers": "계정과목, 차변, 대변, 적요, 거래일자",
                "sample_data": "현금, 1000000, 0, 매출입금, 2025-01-15",
                "domain": "회계관리",
                "confidence": "0.88"
            }
        ]
        
        example_prompt = PromptTemplate(
            input_variables=["headers", "sample_data", "domain", "confidence"],
            template="Headers: {headers}\nSample: {sample_data}\nDomain: {domain}\nConfidence: {confidence}"
        )
        
        few_shot_prompt = FewShotPromptTemplate(
            examples=examples,
            example_prompt=example_prompt,
            prefix="비즈니스 도메인을 분석하는 예시들:\n",
            suffix="Headers: {headers}\nSample: {sample_data}\n\n위 표 데이터의 비즈니스 도메인을 분석하여 JSON으로 응답하세요:\n{{\"domain\": \"도메인명\", \"domain_description\": \"상세설명\", \"confidence\": 0.0-1.0}}",
            input_variables=["headers", "sample_data"]
        )
        
        return LLMChain(
            llm=self.llm,
            prompt=few_shot_prompt,
            output_key="domain_analysis"
        )
    
    def create_semantic_analysis_chain(self) -> LLMChain:
        """Create semantic type analysis chain - preserved from original"""
        
        template = """
도메인: {domain}
컬럼 헤더: {headers}
샘플 데이터: {sample_data}

위 {domain} 도메인의 표 데이터에 대해 각 컬럼의 의미적 타입을 분석하여 JSON으로 응답하세요.

응답 형식:
{{
    "semantic_types": [
        {{
            "name": "의미적_타입명",
            "description": "타입 설명", 
            "patterns": ["정규식패턴들"],
            "confidence": 0.0-1.0,
            "examples": ["예시값들"]
        }}
    ]
}}

컬럼별 의미적 분류를 수행하고, 비즈니스 컨텍스트를 고려한 정확한 타입명을 사용하세요.
"""
        
        prompt = PromptTemplate(
            template=template,
            input_variables=["domain", "headers", "sample_data"]
        )
        
        return LLMChain(
            llm=self.llm,
            prompt=prompt,
            output_key="semantic_types"
        )
    
    def create_relationship_chain(self) -> LLMChain:
        """Create relationship extraction chain - preserved from original"""
        
        template = """
도메인: {domain}
의미적 타입들: {semantic_types}
컬럼 정보: {column_info}

위 {domain} 도메인에서 컬럼들 간의 관계를 분석하여 JSON으로 응답하세요.

관계 유형:
- IDENTIFIES: 식별 관계 (ID → Name)
- CONTAINS: 포함 관계 (Category → Subcategory)  
- REFERENCES: 참조 관계 (Foreign Key)
- DEPENDS_ON: 종속 관계 (Status → Process)
- CATEGORIZES: 분류 관계 (Type → Instance)

응답 형식:
{{
    "relationship_patterns": [
        {{
            "type": "관계유형",
            "from_types": ["소스타입들"],
            "to_types": ["대상타입들"], 
            "description": "관계 설명",
            "confidence": 0.0-1.0
        }}
    ]
}}

비즈니스 로직과 데이터 플로우를 고려하여 의미 있는 관계들을 식별하세요.
"""
        
        prompt = PromptTemplate(
            template=template,
            input_variables=["domain", "semantic_types", "column_info"]
        )
        
        return LLMChain(
            llm=self.llm,
            prompt=prompt,
            output_key="relationships"
        )
    
    async def analyze_sheet_semantics(
        self, 
        headers: List[str], 
        sample_rows: List[Dict[str, Any]]
    ) -> SemanticAnalysis:
        """Perform full semantic analysis on sheet data"""
        
        try:
            # Prepare input data
            headers_str = ", ".join(headers)
            sample_data = self._format_sample_data(sample_rows[:3])
            
            # Step 1: Domain Detection
            domain_chain = self.create_domain_detection_chain()
            domain_result = domain_chain.run({
                "headers": headers_str,
                "sample_data": sample_data
            })
            
            domain_data = self._parse_json_safely(domain_result)
            domain = domain_data.get("domain", "unknown")
            domain_description = domain_data.get("domain_description", "")
            domain_confidence = float(domain_data.get("confidence", 0.5))
            
            # Step 2: Semantic Analysis
            semantic_chain = self.create_semantic_analysis_chain()
            semantic_result = semantic_chain.run({
                "domain": domain,
                "headers": headers_str,
                "sample_data": sample_data
            })
            
            semantic_data = self._parse_json_safely(semantic_result)
            semantic_types = semantic_data.get("semantic_types", [])
            
            # Step 3: Relationship Analysis
            relationship_chain = self.create_relationship_chain()
            relationship_result = relationship_chain.run({
                "domain": domain,
                "semantic_types": json.dumps(semantic_types),
                "column_info": headers_str
            })
            
            relationship_data = self._parse_json_safely(relationship_result)
            relationships = relationship_data.get("relationship_patterns", [])
            
            return SemanticAnalysis(
                domain=domain,
                confidence=domain_confidence,
                semantic_types=semantic_types,
                relationships=relationships,
                analysis_method="langchain"
            )
            
        except Exception as e:
            logger.error(f"Semantic analysis failed: {str(e)}")
            return SemanticAnalysis(
                domain="unknown",
                confidence=0.3,
                semantic_types=[],
                relationships=[],
                analysis_method="fallback"
            )
    
    def _format_sample_data(self, rows: List[Dict[str, Any]]) -> str:
        """Format sample data for prompts"""
        formatted_rows = []
        for i, row in enumerate(rows):
            row_items = [f"{k}={v}" for k, v in row.items() if v][:5]
            formatted_rows.append(f"Row {i+1}: {', '.join(row_items)}")
        return "\n".join(formatted_rows)
    
    def _parse_json_safely(self, json_str: str) -> Dict[str, Any]:
        """Safely parse JSON with error handling"""
        try:
            json_start = json_str.find('{')
            json_end = json_str.rfind('}') + 1
            if json_start >= 0 and json_end > json_start:
                clean_json = json_str[json_start:json_end]
                return json.loads(clean_json)
            return {}
        except Exception:
            return {}


class ExcelProcessor(FileProcessorInterface):
    """Excel file processor implementing sophisticated analysis from original script"""
    
    def __init__(self):
        self.semantic_analyzer = ExcelSemanticAnalyzer()
    
    def get_supported_extensions(self) -> List[str]:
        """Get supported Excel file extensions"""
        return [".xlsx", ".xls"]
    
    def validate_options(self, options: ExcelProcessingOptions) -> bool:
        """Validate Excel processing options"""
        try:
            # Basic validation
            if options.chunk_size <= 0 or options.chunk_overlap < 0:
                return False
            if options.chunk_overlap >= options.chunk_size:
                return False
            return True
        except Exception:
            return False
    
    async def process_file(
        self,
        file_path: str,
        options: ExcelProcessingOptions,
        job_id: str
    ) -> Dict[str, Any]:
        """Process Excel file with comprehensive analysis"""
        
        start_time = datetime.now()
        
        try:
            logger.info(f"Processing Excel file: {os.path.basename(file_path)} (Job: {job_id})")
            
            # Load workbook
            workbook = load_workbook(file_path, read_only=False, data_only=False)
            
            # Process sheets
            sheet_analyses = []
            all_chunks = []
            all_relationships = []
            
            sheets_to_process = self._get_sheets_to_process(workbook, options)
            
            for sheet_name in sheets_to_process:
                sheet = workbook[sheet_name]
                
                if self._is_sheet_empty(sheet):
                    continue
                
                # Analyze sheet
                sheet_analysis = await self._analyze_sheet(sheet, sheet_name, options)
                sheet_analyses.append(sheet_analysis)
                
                # Generate chunks for sheet
                sheet_chunks, sheet_relationships = await self._create_sheet_chunks(
                    sheet, sheet_name, sheet_analysis, options, job_id
                )
                all_chunks.extend(sheet_chunks)
                all_relationships.extend(sheet_relationships)
                
                # Extract additional relationships
                additional_relationships = self._extract_sheet_relationships(
                    sheet_analysis, sheet_name, job_id
                )
                all_relationships.extend(additional_relationships)
            
            workbook.close()
            
            # Create file analysis
            file_analysis = self._create_file_analysis(
                file_path, sheet_analyses, start_time
            )
            
            # Normalize chunk source to use original filename instead of sheet name
            # This ensures Qdrant stores/retrieves by the uploaded file name
            base_filename = os.path.basename(file_path)
            for ch in all_chunks:
                try:
                    ch.source_file = base_filename
                except Exception:
                    pass

            # Generate embeddings if requested
            embeddings = []
            if options.generate_embeddings:
                from app.core.azure_embedding import AzureEmbeddingService
                embedding_service = AzureEmbeddingService()
                await embedding_service.initialize()
                embeddings = await embedding_service.embed_chunks(all_chunks)
            
            end_time = datetime.now()
            processing_duration = (end_time - start_time).total_seconds()
            
            return {
                "job_id": job_id,
                "file_analysis": file_analysis,
                "chunks": all_chunks,
                "embeddings": embeddings,
                "relationships": all_relationships,
                "processing_duration_seconds": processing_duration,
                "options_used": options.dict()
            }
            
        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}")
            raise ProcessingException(f"Excel processing failed: {str(e)}")
    
    async def _analyze_sheet(
        self, 
        sheet: Worksheet, 
        sheet_name: str, 
        options: ExcelProcessingOptions
    ) -> ExcelSheetAnalysis:
        """Analyze single Excel sheet"""
        
        # Extract basic data
        header_row = self._detect_header_row(sheet)
        headers, rows = self._extract_sheet_data(sheet, header_row)
        
        if not headers:
            return ExcelSheetAnalysis(
                sheet_name=sheet_name,
                sheet_index=sheet.index if hasattr(sheet, 'index') else 0,
                total_rows=0,
                total_columns=0,
                data_rows=0,
                header_row=header_row,
                columns=[]
            )
        
        # Analyze columns
        column_analyses = []
        for i, header in enumerate(headers):
            column_analysis = self._analyze_column(header, i, rows)
            column_analyses.append(column_analysis)
        
        # Perform semantic analysis if enabled
        semantic_analysis = None
        if options.use_langchain_analysis and rows:
            semantic_analysis = await self.semantic_analyzer.analyze_sheet_semantics(
                headers, rows[:5]
            )
        
        # Calculate quality metrics
        quality_score = self._calculate_sheet_quality(column_analyses)
        
        return ExcelSheetAnalysis(
            sheet_name=sheet_name,
            sheet_index=sheet.index if hasattr(sheet, 'index') else 0,
            total_rows=sheet.max_row,
            total_columns=len(headers),
            data_rows=len(rows),
            header_row=header_row,
            columns=column_analyses,
            overall_quality_score=quality_score,
            detected_domain=semantic_analysis.domain if semantic_analysis else None,
            domain_confidence=semantic_analysis.confidence if semantic_analysis else None
        )
    
    def _analyze_column(
        self, 
        header: str, 
        index: int, 
        rows: List[Dict[str, Any]]
    ) -> ExcelColumnAnalysis:
        """Analyze individual column"""
        
        # Extract column values
        values = [row.get(header) for row in rows if row.get(header) is not None]
        
        # Calculate statistics
        total_values = len(rows)
        null_count = total_values - len(values)
        unique_values = list(set(str(v) for v in values))
        unique_count = len(unique_values)
        completeness = len(values) / total_values if total_values > 0 else 0.0
        
        # Determine data type
        data_type = self._infer_data_type(values)
        
        # Get sample values
        sample_values = values[:5] if values else []
        
        # Calculate top values
        value_counts = {}
        for value in values:
            value_str = str(value)
            value_counts[value_str] = value_counts.get(value_str, 0) + 1
        
        top_values = [
            {"value": k, "count": v} 
            for k, v in sorted(value_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        ]
        
        return ExcelColumnAnalysis(
            column_name=header,
            column_index=index,
            data_type=data_type,
            total_values=total_values,
            null_count=null_count,
            unique_count=unique_count,
            completeness=completeness,
            sample_values=sample_values,
            top_values=top_values
        )
    
    async def _create_sheet_chunks(
        self,
        sheet: Worksheet,
        sheet_name: str,
        analysis: ExcelSheetAnalysis,
        options: ExcelProcessingOptions,
        job_id: str
    ) -> tuple[List[ProcessedChunk], List[ChunkRelationship]]:
        """Create chunks for sheet based on strategy"""
        
        chunks = []
        relationships = []
        
        # Schema chunk
        schema_chunk = ProcessedChunk(
            chunk_id=f"{job_id}_{sheet_name}_schema",
            content=self._create_schema_content(analysis),
            chunk_type=ExcelChunkType.SCHEMA.value,
            source_file=sheet_name,
            container=options.container,
            metadata={
                "sheet_name": sheet_name,
                "total_rows": analysis.total_rows,
                "total_columns": analysis.total_columns,
                "domain": analysis.detected_domain,
                "quality_score": analysis.overall_quality_score
            }
        )
        chunks.append(schema_chunk)
        
        # Data sample chunks
        if analysis.data_rows > 0:
            headers, rows = self._extract_sheet_data(sheet, analysis.header_row)
            
            # Create data chunks based on strategy
            if options.chunking_strategy == ExcelChunkingStrategy.ROW_BASED:
                data_chunks = self._create_row_based_chunks(
                    headers, rows, sheet_name, job_id, options
                )
            elif options.chunking_strategy == ExcelChunkingStrategy.COLUMN_BASED:
                data_chunks = self._create_column_based_chunks(
                    headers, rows, sheet_name, job_id, options
                )
            else:
                # Fallback to row-based chunks
                data_chunks = self._create_row_based_chunks(
                    headers, rows, sheet_name, job_id, options
                )
            
            chunks.extend(data_chunks)
        
        return chunks, relationships
    
    def _create_schema_content(self, analysis: ExcelSheetAnalysis) -> str:
        """Create schema description content"""
        
        content_parts = [
            f"Excel 시트 '{analysis.sheet_name}' 구조 분석:",
            f"- 총 {analysis.total_rows}행, {analysis.total_columns}열",
            f"- 데이터 행: {analysis.data_rows}개",
            f"- 헤더 위치: {analysis.header_row}행"
        ]
        
        if analysis.detected_domain:
            content_parts.append(f"- 검출된 도메인: {analysis.detected_domain} (신뢰도: {analysis.domain_confidence:.2f})")
        
        content_parts.append("\n컬럼 정보:")
        for col in analysis.columns:
            content_parts.append(
                f"  {col.column_name}: {col.data_type} "
                f"(완성도: {col.completeness:.1%}, 고유값: {col.unique_count}개)"
            )
        
        return "\n".join(content_parts)
    
    def _create_semantic_chunks(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        sheet_name: str,
        job_id: str,
        options: ExcelProcessingOptions
    ) -> List[ProcessedChunk]:
        """Create semantically meaningful chunks"""
        
        chunks = []
        
        # Statistical summary chunk
        stats_content = self._create_statistical_summary(headers, rows)
        stats_chunk = ProcessedChunk(
            chunk_id=f"{job_id}_{sheet_name}_statistics",
            content=stats_content,
            chunk_type=ExcelChunkType.STATISTICAL.value,
            source_file=sheet_name,
            container=options.container,
            metadata={"chunk_strategy": "semantic", "content_type": "statistics"}
        )
        chunks.append(stats_chunk)
        
        # Representative data chunks
        chunk_size = min(options.chunk_size // 100, len(rows))  # Adjust for Excel data
        for i in range(0, len(rows), chunk_size):
            chunk_rows = rows[i:i + chunk_size]
            
            data_content = self._create_data_content(headers, chunk_rows, i)
            data_chunk = ProcessedChunk(
                chunk_id=f"{job_id}_{sheet_name}_data_{i}",
                content=data_content,
                chunk_type=ExcelChunkType.DATA_SAMPLE.value,
                source_file=sheet_name,
                container=options.container,
                start_position=i,
                end_position=min(i + chunk_size, len(rows)),
                metadata={
                    "chunk_strategy": "semantic",
                    "row_start": i,
                    "row_count": len(chunk_rows)
                }
            )
            chunks.append(data_chunk)
        
        return chunks
    
    def _create_fixed_chunks(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        sheet_name: str,
        job_id: str,
        options: ExcelProcessingOptions
    ) -> List[ProcessedChunk]:
        """Create fixed-size chunks"""
        
        chunks = []
        chunk_size = options.chunk_size // 50  # Adjust for Excel rows
        
        for i in range(0, len(rows), chunk_size):
            chunk_rows = rows[i:i + chunk_size]
            
            content = self._create_data_content(headers, chunk_rows, i)
            chunk = ProcessedChunk(
                chunk_id=f"{job_id}_{sheet_name}_chunk_{i}",
                content=content,
                chunk_type=ExcelChunkType.DATA_SAMPLE.value,
                source_file=sheet_name,
                container=options.container,
                start_position=i,
                end_position=min(i + chunk_size, len(rows)),
                metadata={
                    "chunk_strategy": "fixed",
                    "chunk_size": chunk_size,
                    "row_count": len(chunk_rows)
                }
            )
            chunks.append(chunk)
        
        return chunks
    
    def _create_row_based_chunks(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        sheet_name: str,
        job_id: str,
        options: ExcelProcessingOptions
    ) -> List[ProcessedChunk]:
        """Create chunks by grouping rows with overlap"""
        
        chunks = []
        if not rows:
            return chunks
        
        chunk_size = options.chunk_size
        overlap = options.chunk_overlap
        step = max(1, chunk_size - overlap)
        
        for i in range(0, len(rows), step):
            chunk_end = min(i + chunk_size, len(rows))
            chunk_rows = rows[i:chunk_end]
            
            if chunk_rows:
                content = self._create_data_content(headers, chunk_rows, i)
                chunk = ProcessedChunk(
                    chunk_id=f"{job_id}_{sheet_name}_row_{i}",
                    content=content,
                    chunk_type=ExcelChunkType.DATA_SAMPLE.value,
                    source_file=sheet_name,
                    container=options.container,
                    start_position=i,
                    end_position=chunk_end,
                    metadata={
                        "chunk_strategy": "row_based",
                        "row_start": i,
                        "row_end": chunk_end,
                        "row_count": len(chunk_rows)
                    }
                )
                chunks.append(chunk)
        
        return chunks
    
    def _create_column_based_chunks(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        sheet_name: str,
        job_id: str,
        options: ExcelProcessingOptions
    ) -> List[ProcessedChunk]:
        """Create chunks by grouping columns with overlap"""
        
        chunks = []
        if not headers or not rows:
            return chunks
        
        chunk_size = options.chunk_size
        overlap = options.chunk_overlap
        step = max(1, chunk_size - overlap)
        
        for i in range(0, len(headers), step):
            chunk_end = min(i + chunk_size, len(headers))
            chunk_headers = headers[i:chunk_end]
            
            # Create content with only selected columns
            content_parts = [f"열 기반 청크 (열 {i + 1}-{chunk_end}):"]
            content_parts.append(f"헤더: {', '.join(chunk_headers)}")
            content_parts.append("데이터:")
            
            for row_idx, row in enumerate(rows[:10]):  # Show first 10 rows
                row_items = [f"{chunk_headers[j]}={row.get(chunk_headers[j])}" 
                            for j in range(len(chunk_headers)) 
                            if row.get(chunk_headers[j]) is not None]
                if row_items:
                    content_parts.append(f"  행 {row_idx + 1}: {', '.join(row_items)}")
            
            if len(rows) > 10:
                content_parts.append(f"  ... 총 {len(rows)}행 포함")
            
            content = "\n".join(content_parts)
            
            chunk = ProcessedChunk(
                chunk_id=f"{job_id}_{sheet_name}_col_{i}",
                content=content,
                chunk_type=ExcelChunkType.DATA_SAMPLE.value,
                source_file=sheet_name,
                container=options.container,
                start_position=i,
                end_position=chunk_end,
                metadata={
                    "chunk_strategy": "column_based",
                    "column_start": i,
                    "column_end": chunk_end,
                    "column_count": len(chunk_headers)
                }
            )
            chunks.append(chunk)
        
        return chunks
    
    def _create_statistical_summary(self, headers: List[str], rows: List[Dict[str, Any]]) -> str:
        """Create statistical summary content"""
        
        summary_parts = [f"데이터 통계 요약 (총 {len(rows)}행):"]
        
        for header in headers[:10]:  # Limit to first 10 columns
            values = [row.get(header) for row in rows if row.get(header) is not None]
            if values:
                unique_count = len(set(str(v) for v in values))
                completeness = len(values) / len(rows)
                
                summary_parts.append(
                    f"  {header}: {len(values)}개 값, {unique_count}개 고유값, "
                    f"완성도 {completeness:.1%}"
                )
        
        return "\n".join(summary_parts)
    
    def _create_data_content(
        self, 
        headers: List[str], 
        rows: List[Dict[str, Any]], 
        start_idx: int
    ) -> str:
        """Create content for data chunks"""
        
        content_parts = [f"데이터 샘플 (시작: {start_idx + 1}행):"]
        
        for i, row in enumerate(rows[:5]):  # Show first 5 rows of chunk
            row_items = [f"{k}={v}" for k, v in row.items() if v is not None][:8]
            content_parts.append(f"  행 {start_idx + i + 1}: {', '.join(row_items)}")
        
        if len(rows) > 5:
            content_parts.append(f"  ... 총 {len(rows)}행 포함")
        
        return "\n".join(content_parts)
    
    def _extract_sheet_relationships(
        self,
        analysis: ExcelSheetAnalysis,
        sheet_name: str,
        job_id: str
    ) -> List[ChunkRelationship]:
        """Extract relationships from sheet analysis"""
        
        relationships = []
        
        # Create relationships between schema and data chunks
        schema_chunk_id = f"{job_id}_{sheet_name}_schema"
        
        # Find data chunks for this sheet
        for i in range(0, analysis.data_rows, 50):  # Assuming chunk size
            data_chunk_id = f"{job_id}_{sheet_name}_data_{i}"
            
            relationship = ChunkRelationship(
                relationship_id=f"{job_id}_{sheet_name}_schema_data_{i}",
                from_chunk_id=schema_chunk_id,
                to_chunk_id=data_chunk_id,
                relationship_type="contains",
                confidence=0.9,
                description=f"Schema contains data chunk {i}"
            )
            relationships.append(relationship)
        
        return relationships
    
    # Helper methods from original script
    def _get_sheets_to_process(
        self, 
        workbook, 
        options: ExcelProcessingOptions
    ) -> List[str]:
        """Get list of sheets to process"""
        
        all_sheets = workbook.sheetnames
        
        if options.sheet_filter:
            # Filter to specified sheets
            sheets = [s for s in options.sheet_filter if s in all_sheets]
        else:
            sheets = all_sheets
        
        if options.skip_empty_sheets:
            non_empty_sheets = []
            for sheet_name in sheets:
                sheet = workbook[sheet_name]
                if not self._is_sheet_empty(sheet):
                    non_empty_sheets.append(sheet_name)
            sheets = non_empty_sheets
        
        return sheets
    
    def _is_sheet_empty(self, sheet: Worksheet) -> bool:
        """Check if sheet is empty"""
        return sheet.max_row <= 1 and sheet.max_column <= 1
    
    def _detect_header_row(self, sheet: Worksheet) -> int:
        """Detect header row - preserved from original script"""
        
        for row_num in range(1, min(6, sheet.max_row + 1)):
            row_cells = []
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell = sheet.cell(row=row_num, column=col)
                if cell.value is not None:
                    row_cells.append(str(cell.value))
            
            if len(row_cells) >= 3:
                if self._calculate_header_score(row_cells) > 0.6:
                    return row_num
        return 1
    
    def _calculate_header_score(self, cells: List[str]) -> float:
        """Calculate header score - preserved from original script"""
        
        import re
        score = 0.0
        total_cells = len(cells)
        
        for cell_value in cells:
            cell_value = cell_value.strip()
            if not cell_value:
                continue
                
            if re.search(r'[가-힣]', cell_value):
                score += 0.8
            elif re.search(r'^[A-Za-z][A-Za-z\s_-]*$', cell_value):
                score += 0.7
            elif re.search(r'[A-Za-z]', cell_value) and len(cell_value) < 30:
                score += 0.5
            elif re.search(r'^\d+$', cell_value):
                score -= 0.3
            elif len(cell_value) > 50:
                score -= 0.2
        
        return score / total_cells if total_cells > 0 else 0.0
    
    def _extract_sheet_data(
        self, 
        sheet: Worksheet, 
        header_row: int
    ) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Extract data from sheet - preserved from original script"""
        
        max_col = self._find_last_data_column(sheet, header_row)
        
        headers = []
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=header_row, column=col)
            header = str(cell.value) if cell.value is not None else f"Column_{col}"
            headers.append(header)
        
        rows = []
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for i, header in enumerate(headers):
                cell = sheet.cell(row=row_num, column=i + 1)
                value = cell.value
                
                if value is not None:
                    has_data = True
                    row_data[header] = value
                else:
                    row_data[header] = None
            
            if has_data:
                rows.append(row_data)
        
        return headers, rows
    
    def _find_last_data_column(self, sheet: Worksheet, header_row: int) -> int:
        """Find last column with data"""
        max_col = 0
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row, column=col)
            if cell.value is not None:
                max_col = col
        return max_col
    
    def _infer_data_type(self, values: List[Any]) -> str:
        """Infer data type from values"""
        if not values:
            return "empty"
        
        # Check for numeric
        numeric_count = 0
        date_count = 0
        
        for value in values[:20]:  # Sample first 20 values
            if isinstance(value, (int, float)):
                numeric_count += 1
            elif isinstance(value, datetime):
                date_count += 1
        
        if numeric_count > len(values) * 0.8:
            return "numeric"
        elif date_count > len(values) * 0.5:
            return "date"
        else:
            return "text"
    
    def _calculate_sheet_quality(self, columns: List[ExcelColumnAnalysis]) -> float:
        """Calculate overall sheet quality score"""
        if not columns:
            return 0.0
        
        completeness_scores = [col.completeness for col in columns]
        return sum(completeness_scores) / len(completeness_scores)
    
    def _create_file_analysis(
        self,
        file_path: str,
        sheet_analyses: List[ExcelSheetAnalysis],
        start_time: datetime
    ) -> ExcelFileAnalysis:
        """Create complete file analysis"""
        
        total_data_points = sum(
            sheet.total_rows * sheet.total_columns 
            for sheet in sheet_analyses
        )
        
        overall_quality = sum(
            sheet.overall_quality_score or 0.0 
            for sheet in sheet_analyses
        ) / len(sheet_analyses) if sheet_analyses else 0.0
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        return ExcelFileAnalysis(
            filename=os.path.basename(file_path),
            file_size=os.path.getsize(file_path),
            total_sheets=len(sheet_analyses),
            processed_sheets=len(sheet_analyses),
            sheets=sheet_analyses,
            overall_quality_score=overall_quality,
            total_data_points=total_data_points,
            total_relationships=0,  # Will be calculated later
            processing_time_seconds=processing_time,
            langchain_analysis_used=True
        )